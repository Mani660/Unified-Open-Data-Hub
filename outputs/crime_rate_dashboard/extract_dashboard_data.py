import json
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl

INPUT = Path(r"C:\Users\Manish\OneDrive\Desktop\All_Data_Consolidated.xlsx")
OUTPUT = Path(r"C:\Users\Manish\OneDrive\ドキュメント\New project\outputs\crime_rate_dashboard\dashboard_data.js")


def clean(value):
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return value


def parse_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    text = str(value).strip()
    for fmt in ("%d-%m-%Y %H:%M", "%d-%m-%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    return None


def top(counter, n=12):
    return [{"name": k, "value": v} for k, v in counter.most_common(n) if k not in (None, "None")]


def pct(part, total):
    return 0 if not total else round((part / total) * 100, 2)


wb = openpyxl.load_workbook(INPUT, read_only=True, data_only=True)
ws = wb["Unified_Dataset"]
headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
idx = {name: i for i, name in enumerate(headers)}

sources = Counter()

india_total = 0
india_closed = 0
police_total = 0
police_count = 0
age_total = 0
age_count = 0
city_counts = Counter()
city_closed = Counter()
domain_counts = Counter()
weapon_counts = Counter()
gender_counts = Counter()
hour_counts = Counter()
month_counts = Counter()
city_domain = defaultdict(Counter)

world_cities = []
juvenile_trend = []
violent_region = Counter()
violent_indicator = Counter()
violent_year = Counter()
violent_country = Counter()
drug_region = Counter()
drug_offence = Counter()
drug_year = Counter()

for row in ws.iter_rows(min_row=2, values_only=True):
    source = clean(row[idx["Dataset_Source"]])
    sources[source] += 1

    if source == "India_Crime_Dataset":
        india_total += 1
        city = clean(row[idx["City"]]) or "Unknown"
        domain = clean(row[idx["Crime Domain"]]) or "Unknown"
        weapon = clean(row[idx["Weapon Used"]]) or "Unknown"
        gender = clean(row[idx["Victim Gender"]]) or "Unknown"
        closed = clean(row[idx["Case Closed"]])
        police = row[idx["Police Deployed"]]
        age = row[idx["Victim Age"]]
        occurrence = parse_date(row[idx["Date of Occurrence"]])
        time_value = parse_date(row[idx["Time of Occurrence"]])

        city_counts[city] += 1
        domain_counts[domain] += 1
        weapon_counts[weapon] += 1
        gender_counts[gender] += 1
        city_domain[city][domain] += 1
        if closed == "Yes":
            india_closed += 1
            city_closed[city] += 1
        if isinstance(police, (int, float)):
            police_total += police
            police_count += 1
        if isinstance(age, (int, float)):
            age_total += age
            age_count += 1
        if occurrence:
            month_counts[occurrence.strftime("%b %Y")] += 1
        if time_value:
            hour_counts[f"{time_value.hour:02d}:00"] += 1

    elif source == "World_Crime_Index":
        city = clean(row[idx["City"]])
        crime = row[idx["Crime Index"]]
        safety = row[idx["Safety Index"]]
        rank = row[idx["Rank"]]
        if city and isinstance(crime, (int, float)) and isinstance(safety, (int, float)):
            world_cities.append({"city": city, "rank": rank, "crimeIndex": round(crime, 2), "safetyIndex": round(safety, 2)})

    elif source == "India_Juvenile_Crime_Hist":
        year = row[idx["Year"]]
        rate = row[idx["Volume of juvenile crime per lakh of population"]]
        total_cases = row[idx["Total cognizable crime cases under IPC"]]
        juvenile_cases = row[idx["Total juvenile crime cases under IPC"]]
        if isinstance(year, (int, float)):
            juvenile_trend.append({
                "year": int(year),
                "ratePerLakh": rate,
                "totalCases": total_cases,
                "juvenileCases": juvenile_cases,
                "juvenileShare": row[idx["Percentage of juvenile crime to total cognizable crime"]],
            })

    elif source == "Violent_Sexual_Crimes_Global":
        value = row[idx["VALUE"]]
        region = clean(row[idx["Region"]])
        indicator = clean(row[idx["Indicator"]])
        country = clean(row[idx["Country"]])
        year = row[idx["Year"]]
        if isinstance(value, (int, float)):
            violent_region[region] += value
            violent_indicator[indicator] += value
            violent_country[country] += value
            if isinstance(year, (int, float)):
                violent_year[int(year)] += value

    elif source == "Drug_Related_Offences":
        value = row[idx["Calculated total"]]
        region = clean(row[idx["Region"]])
        offence = clean(row[idx["Type of offence"]])
        year = row[idx["Year"]]
        if isinstance(value, (int, float)):
            drug_region[region] += value
            drug_offence[offence] += value
            if isinstance(year, (int, float)):
                drug_year[int(year)] += value

city_cards = []
for city, count in city_counts.most_common(15):
    city_cards.append({
        "city": city,
        "cases": count,
        "closureRate": pct(city_closed[city], count),
        "topDomain": city_domain[city].most_common(1)[0][0] if city_domain[city] else "Unknown",
    })

dashboard = {
    "meta": {
        "rows": sum(sources.values()),
        "sources": top(sources, 20),
    },
    "india": {
        "totalIncidents": india_total,
        "closureRate": pct(india_closed, india_total),
        "avgPolice": round(police_total / police_count, 1) if police_count else 0,
        "avgVictimAge": round(age_total / age_count, 1) if age_count else 0,
        "cities": city_cards,
        "domains": top(domain_counts, 10),
        "weapons": top(weapon_counts, 10),
        "gender": top(gender_counts, 8),
        "hours": [{"name": k, "value": hour_counts[k]} for k in sorted(hour_counts)],
        "months": [{"name": k, "value": v} for k, v in month_counts.items()],
    },
    "worldCrimeIndex": {
        "cities": sorted(world_cities, key=lambda item: item["rank"])[:80],
        "avgCrimeIndex": round(sum(item["crimeIndex"] for item in world_cities) / len(world_cities), 2) if world_cities else 0,
        "avgSafetyIndex": round(sum(item["safetyIndex"] for item in world_cities) / len(world_cities), 2) if world_cities else 0,
    },
    "juvenile": sorted(juvenile_trend, key=lambda item: item["year"]),
    "globalViolent": {
        "byRegion": top(violent_region, 8),
        "byIndicator": top(violent_indicator, 8),
        "byCountry": top(violent_country, 12),
        "byYear": [{"name": str(k), "value": violent_year[k]} for k in sorted(violent_year)],
    },
    "drugOffences": {
        "byRegion": top(drug_region, 8),
        "byOffence": top(drug_offence, 8),
        "byYear": [{"name": str(k), "value": drug_year[k]} for k in sorted(drug_year)],
    },
}

OUTPUT.write_text("window.CRIME_DASHBOARD_DATA = " + json.dumps(dashboard, ensure_ascii=False) + ";\n", encoding="utf-8")
print(json.dumps({
    "output": str(OUTPUT),
    "rows": dashboard["meta"]["rows"],
    "india_incidents": dashboard["india"]["totalIncidents"],
    "world_index_cities": len(dashboard["worldCrimeIndex"]["cities"]),
}, indent=2))
